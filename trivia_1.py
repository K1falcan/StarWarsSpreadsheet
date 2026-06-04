import random
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Styles (matching existing tabs) ──
gold_fill   = PatternFill("solid", start_color="FFD700", end_color="FFD700")
dark_fill   = PatternFill("solid", start_color="1A1A2E", end_color="1A1A2E")
alt_fill    = PatternFill("solid", start_color="2E2E4E", end_color="2E2E4E")
green_fill  = PatternFill("solid", start_color="1A3A1A", end_color="1A3A1A")
red_fill    = PatternFill("solid", start_color="3A1A1A", end_color="3A1A1A")
title_font  = Font(name="Arial", bold=True, color="FFFFFF", size=14)
header_font = Font(name="Arial", bold=True, color="1A1A2E", size=12)
data_font   = Font(name="Arial", color="FFFFFF", size=11)
answer_font = Font(name="Arial", color="FFD700", size=11, bold=True)
center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="FFD700"),
    right=Side(style="thin", color="FFD700"),
    top=Side(style="thin", color="FFD700"),
    bottom=Side(style="thin", color="FFD700"),
)

def style_header(cell):
    cell.fill = gold_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = thin_border

def style_data(cell, row_idx, align=center):
    cell.fill = dark_fill if row_idx % 2 == 0 else alt_fill
    cell.font = data_font
    cell.alignment = align
    cell.border = thin_border

def style_answer(cell, row_idx):
    cell.fill = dark_fill if row_idx % 2 == 0 else alt_fill
    cell.font = answer_font
    cell.alignment = left
    cell.border = thin_border

# ── Question bank (drawn from spreadsheet data) ──
ALL_QUESTIONS = [
    # Movie order
    {
        "category": "Movie Order",
        "question": "Which film is #1 in chronological order?",
        "a": "The Phantom Menace", "b": "A New Hope",
        "c": "Attack of the Clones", "d": "Revenge of the Sith",
        "answer": "A",
    },
    {
        "category": "Movie Order",
        "question": "Which Star Wars film was released first in theaters?",
        "a": "The Phantom Menace", "b": "The Empire Strikes Back",
        "c": "A New Hope", "d": "Return of the Jedi",
        "answer": "C",
    },
    {
        "category": "Movie Order",
        "question": "What is the chronological position of 'Return of the Jedi'?",
        "a": "4th", "b": "5th", "c": "7th", "d": "6th",
        "answer": "D",
    },
    {
        "category": "Movie Order",
        "question": "Which film comes immediately before 'The Force Awakens' in chronological order?",
        "a": "Return of the Jedi", "b": "The Last Jedi",
        "c": "Revenge of the Sith", "d": "A New Hope",
        "answer": "A",
    },
    {
        "category": "Movie Order",
        "question": "In release order, which film was released 4th?",
        "a": "A New Hope", "b": "Return of the Jedi",
        "c": "The Phantom Menace", "d": "Attack of the Clones",
        "answer": "C",
    },
    {
        "category": "Movie Order",
        "question": "How many Star Wars main saga films are listed in the spreadsheet?",
        "a": "6", "b": "7", "c": "9", "d": "8",
        "answer": "C",
    },
    # Starships
    {
        "category": "Starships",
        "question": "Which starship has a hyperdrive rating of 0.5 (the fastest)?",
        "a": "X-wing", "b": "Millennium Falcon",
        "c": "Jedi Interceptor", "d": "A-wing",
        "answer": "B",
    },
    {
        "category": "Starships",
        "question": "What is the length of the Death Star in meters?",
        "a": "1,600 m", "b": "19,000 m", "c": "120,000 m", "d": "3,170 m",
        "answer": "C",
    },
    {
        "category": "Starships",
        "question": "Which starship has the highest atmospheric speed in the dataset (1,500 km/h)?",
        "a": "A-wing", "b": "TIE Advanced x1",
        "c": "Jedi Interceptor", "d": "Naboo fighter",
        "answer": "C",
    },
    {
        "category": "Starships",
        "question": "Who manufactures the Millennium Falcon?",
        "a": "Kuat Drive Yards", "b": "Sienar Fleet Systems",
        "c": "Incom Corporation", "d": "Corellian Engineering Corporation",
        "answer": "D",
    },
    {
        "category": "Starships",
        "question": "What class of ship is the Slave 1?",
        "a": "Starfighter", "b": "Patrol craft",
        "c": "Light freighter", "d": "Assault Starfighter",
        "answer": "B",
    },
    {
        "category": "Starships",
        "question": "The Executor is 19,000 meters long. What class is it?",
        "a": "Star Destroyer", "b": "Deep Space Battlestation",
        "c": "Star dreadnought", "d": "Capital ship",
        "answer": "C",
    },
    # Species
    {
        "category": "Species",
        "question": "Which species has an average lifespan of 1,000 years?",
        "a": "Wookie", "b": "Yoda's species", "c": "Human", "d": "Hutt",
        "answer": "D",
    },
    {
        "category": "Species",
        "question": "What language do Wookies speak?",
        "a": "Huttese", "b": "Shyriiwook",
        "c": "Galactic Basic", "d": "Ewokese",
        "answer": "B",
    },
    {
        "category": "Species",
        "question": "What is the average height of Yoda's species (in cm)?",
        "a": "100 cm", "b": "80 cm", "c": "66 cm", "d": "120 cm",
        "answer": "C",
    },
    {
        "category": "Species",
        "question": "Which species is classified as a gastropod?",
        "a": "Rodian", "b": "Gungan", "c": "Hutt", "d": "Trandoshan",
        "answer": "C",
    },
    {
        "category": "Species",
        "question": "What is the average lifespan of a Wookie?",
        "a": "120 years", "b": "400 years", "c": "900 years", "d": "250 years",
        "answer": "B",
    },
    {
        "category": "Species",
        "question": "What language do Ewoks speak?",
        "a": "Sullutese", "b": "Ewokese", "c": "Huttese", "d": "Dosh",
        "answer": "B",
    },
]

def add_trivia_tab(filepath):
    wb = load_workbook(filepath)

    if "Trivia" in wb.sheetnames:
        del wb["Trivia"]

    ws = wb.create_sheet(title="Trivia")

    questions = random.sample(ALL_QUESTIONS, 10)

    # ── Column widths ──
    ws.column_dimensions["A"].width = 5   # #
    ws.column_dimensions["B"].width = 18  # Category
    ws.column_dimensions["C"].width = 42  # Question
    ws.column_dimensions["D"].width = 30  # A
    ws.column_dimensions["E"].width = 30  # B
    ws.column_dimensions["F"].width = 30  # C
    ws.column_dimensions["G"].width = 30  # D
    ws.column_dimensions["H"].width = 10  # Answer

    ws.sheet_view.showGridLines = False

    # ── Title row ──
    ws.merge_cells("A1:H1")
    ws["A1"] = "Star Wars Trivia — 10 Questions"
    ws["A1"].fill = PatternFill("solid", start_color="4A0000", end_color="4A0000")
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].border = thin_border
    ws.row_dimensions[1].height = 28

    # ── Header row ──
    headers = ["#", "Category", "Question", "A", "B", "C", "D", "Answer"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        style_header(cell)
    ws.row_dimensions[2].height = 22

    # ── Question rows ──
    for i, q in enumerate(questions):
        row = i + 3
        ws.row_dimensions[row].height = 52

        vals = [
            (1, i + 1, center),
            (2, q["category"], center),
            (3, q["question"], left),
            (4, q["a"], left),
            (5, q["b"], left),
            (6, q["c"], left),
            (7, q["d"], left),
        ]
        for col, val, align in vals:
            cell = ws.cell(row=row, column=col, value=val)
            style_data(cell, i, align)

        answer_cell = ws.cell(row=row, column=8, value=q["answer"])
        style_answer(answer_cell, i)

    wb.save(filepath)
    print(f"Trivia tab added with 10 questions → {filepath}")

if __name__ == "__main__":
    add_trivia_tab("star_wars_order.xlsx")
